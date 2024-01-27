import os
from typing import List, Dict, cast
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill
from src.api.base_api import BaseApi


classes = os.getenv('CLASS_PATH')
assert classes is not None, "CLASS_PATH is None"
strings = os.getenv('STRING_PATH')
assert strings, "STRING_PATH is None"
assert strings, "STRING_PATH is None"
adds = os.getenv('ADDS')


class Checker:
    def __init__(self,
                 api: BaseApi
                 ) -> None:
        self.api = api

    def convert_exc_to_dict(self) -> Dict[int, str]:
        print(f"run load {classes}")
        assert classes is not None, "CLASS_PATH is None"
        workbook = load_workbook(classes)
        print(f"load {classes} success")
        sheet = workbook.active
        sheet = cast(Worksheet, workbook.active)
        assert sheet, "sheet is None"
        data_dict: dict = {}

        for row in sheet.iter_rows(  # type: ignore
            min_row=2,
            values_only=True
        ):
            key = row[0]
            value = str(row[1]) if row[1] is not None else ""
            data_dict[key] = value
        return data_dict

    def create_prompt(self,
                      line1: str,
                      line2: str,
                      adds: str = ''
                      ) -> List[Dict[str, str]]:

        print('__________________')
        print(f'{line1} - класс')
        print(f'{line2} - предложение')
        return [
            {
                'role': 'system',
                'content': (
                    'Ты модель классификатор текста. '
                    'Тебе дается тематика предложения и само предложение. '
                    'Ты выявляешь наличие синонимичного смысла между заданной темой и предложением.  '
                    f'Ты определяешь, является ли текст "{line2}" и тематика '
                    f'"{line1}" одинаковыми по смыслу. Ты учитываешь, что субъект, '
                    'на которого направлен акцент, в теме и в предложении должен быть один и тот же. '
                    'Если есть какие-то противоречия, то ты отвечаешь Нет. '
                    'Ты отвечаешь только Да или Нет, и говоришь почему '
                    f'{adds}'
                )
            },
            {
                'role': 'user',
                'content': (
                    'предложение 1: "в какой посуде можно приготовить блюдо?"\n'
                    'предложение 2: "борщ лучше варить в кастрюле?"'
                )
            },
            {
                'role': 'assistant',
                'content': 'Да, потому что говорится о возможности '
                           'использования посуды для приготовления блюда борщ.'

            },
            {
                'role': 'user',
                'content': (
                    'предложение 1: "в какой посуде можно приготовить блюдо?"\n'
                    'предложение 2: "зачем мне кастрюля?"'
                )
            },
            {
                'role': 'assistant',
                'content': 'Нет, потому что класс предполагает допустимую '
                           'посуду, а предложение выявляет потребность наличия посуды.'

            },
            {
                'role': 'user',
                'content': (
                    f'предложение 1: "{line1}"\n'
                    f'предложение 2: "{line2}"'
                )
            }
        ]

    def run(self) -> None:
        data_dict = self.convert_exc_to_dict()
        # Открытие эксель файла с работой дата-инженера
        print(f"run load {strings}")
        assert strings, "STRING_PATH is None"
        workbook = load_workbook(
            strings,
            data_only=True
        )
        print(f"load {strings} success")
        assert workbook.active, "sheet is None"
        sheet = cast(Worksheet, workbook.active)

        current_row = 2
        # Перебор строк в рабочем Excel файле

        while current_row <= sheet.max_row:

            tagged_id = sheet.cell(
                row=current_row,
                column=1
            ).value
            if tagged_id is None:
                continue
            if isinstance(tagged_id, (float, str)):
                itagged_id = int(tagged_id)
            elif isinstance(tagged_id, int):
                itagged_id = int(tagged_id)
            else:
                continue
            correct_line = data_dict.get(itagged_id)
            checked_line = sheet.cell(
                row=current_row,
                column=2
            ).value
            if (
                (
                    not checked_line or
                    not isinstance(checked_line, str)
                ) or
                (
                    not correct_line or
                    not isinstance(correct_line, str)
                )
            ):
                continue
            rsp = self.api.completions(
                messages=self.create_prompt(
                    correct_line,
                    checked_line,
                    adds=''
                )
            )
            print(rsp)
            if 'нет' in rsp.lower():
                cell = sheet.cell(
                    row=current_row,
                    column=2
                )
                cell.fill = PatternFill(
                    start_color="FF0000",
                    end_color="FF0000",
                    fill_type='solid'
                )
            elif 'да' in rsp.lower():
                cell = sheet.cell(
                    row=current_row,
                    column=2
                )
                cell.fill = PatternFill(
                    start_color="FFFFFF",
                    end_color="FFFFFF",
                    fill_type='solid'
                )
            else:
                cell = sheet.cell(
                    row=current_row,
                    column=2
                )
                cell.fill = PatternFill(
                    start_color="FFFF00",
                    end_color="FFFF00",
                    fill_type='solid'
                )

            current_row += 1
        workbook.save(strings)
